"""Tests for src/export.py.

Covers unit tests for every public function and both private helpers
(_drop_db_cols, _sheet_safe).  The two openpyxl-mutation helpers
(_autosize_columns, _freeze_header) are exercised indirectly through
write_workbook integration tests that inspect the resulting .xlsx file.
Integration-level tests use tmp_path directories with real SQLite databases
and real .xlsx file I/O.
"""

import sqlite3
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

# Make src/ importable without installing the package
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from export import (
    ExportResult,
    _drop_db_cols,
    _sheet_safe,
    build_sheet_map,
    build_summary,
    export,
    read_consolidated,
    read_quarantine,
    split_by_column,
    write_workbook,
)


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------


def _make_clean_df(n: int = 3, region: str = "East") -> pd.DataFrame:
    """Return a minimal clean DataFrame matching the consolidated table schema (no id)."""
    return pd.DataFrame({
        "source_file": ["sales.xlsx"] * n,
        "source_row":  list(range(2, n + 2)),
        "date":        ["2024-01-15"] * n,
        "product":     ["Widget A"] * n,
        "region":      [region] * n,
        "sales_rep":   ["Alice"] * n,
        "customer":    ["Acme Corp"] * n,
        "quantity":    [10.0] * n,
        "revenue":     [1000.0] * n,
        "loaded_at":   ["2024-01-15T10:00:00Z"] * n,
    })


def _make_quarantine_df(n: int = 1) -> pd.DataFrame:
    """Return a minimal quarantine DataFrame (no id)."""
    return pd.DataFrame({
        "quarantine_reason": ["revenue is negative (-50) in row 4 of sales.xlsx"] * n,
        "source_file": ["sales.xlsx"] * n,
        "source_row":  list(range(4, n + 4)),
        "date":        ["2024-03-01"] * n,
        "product":     ["Widget C"] * n,
        "region":      ["North"] * n,
        "sales_rep":   ["Bob"] * n,
        "customer":    ["Initech"] * n,
        "quantity":    ["3.0"] * n,
        "revenue":     ["-50.0"] * n,
        "quarantined_at": ["2024-01-15T10:00:00Z"] * n,
    })


def _make_db(
    tmp_path: Path,
    n_clean: int = 3,
    n_quarantine: int = 1,
    region: str = "East",
) -> Path:
    """Create a minimal SQLite database with consolidated and quarantine tables.

    Args:
        tmp_path:     pytest tmp_path directory.
        n_clean:      Number of rows to insert into the consolidated table.
        n_quarantine: Number of rows to insert into the quarantine table.
        region:       Region value to set on every consolidated row.

    Returns:
        Path to the created SQLite file.
    """
    db = tmp_path / "test.db"
    with sqlite3.connect(db) as conn:
        conn.executescript("""
            CREATE TABLE consolidated (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file TEXT,
                source_row  INTEGER,
                date        TEXT,
                product     TEXT,
                region      TEXT,
                sales_rep   TEXT,
                customer    TEXT,
                quantity    REAL,
                revenue     REAL,
                loaded_at   TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ', 'now'))
            );
            CREATE TABLE quarantine (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                quarantine_reason TEXT NOT NULL,
                source_file       TEXT,
                source_row        INTEGER,
                date              TEXT,
                product           TEXT,
                region            TEXT,
                sales_rep         TEXT,
                customer          TEXT,
                quantity          TEXT,
                revenue           TEXT,
                quarantined_at    TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ', 'now'))
            );
        """)
        for i in range(n_clean):
            conn.execute(
                "INSERT INTO consolidated "
                "(source_file, source_row, date, product, region, "
                "sales_rep, customer, quantity, revenue) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("sales.xlsx", i + 2, "2024-01-15", "Widget A",
                 region, "Alice", "Acme Corp", 10.0, 1000.0),
            )
        for i in range(n_quarantine):
            conn.execute(
                "INSERT INTO quarantine "
                "(quarantine_reason, source_file, source_row, date, product, "
                "region, sales_rep, customer, quantity, revenue) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("revenue is negative (-50) in row 10 of sales.xlsx",
                 "sales.xlsx", i + 10, "2024-02-01", "Widget B",
                 "North", "Bob", "Corp Inc", "3.0", "-50.0"),
            )
        conn.commit()
    return db


def _make_result(
    output_path: Path | None = None,
    sheets: list[str] | None = None,
    n_clean: int = 10,
    n_quarantine: int = 2,
) -> ExportResult:
    """Return an ExportResult with sensible defaults."""
    return ExportResult(
        output_path=output_path or Path("data/output/out.xlsx"),
        sheets=sheets or ["Clean Data", "Quarantine"],
        n_clean=n_clean,
        n_quarantine=n_quarantine,
    )


# ---------------------------------------------------------------------------
# _drop_db_cols
# ---------------------------------------------------------------------------


class TestDropDbCols:
    """Tests for the private _drop_db_cols helper."""

    def test_id_column_removed_when_present(self) -> None:
        """The 'id' column is removed from the DataFrame."""
        df = pd.DataFrame({"id": [1, 2], "revenue": [100.0, 200.0]})
        result = _drop_db_cols(df)
        assert "id" not in result.columns

    def test_other_columns_retained(self) -> None:
        """Columns other than the drop set are left in place."""
        df = pd.DataFrame({"id": [1], "revenue": [100.0], "product": ["Widget"]})
        result = _drop_db_cols(df)
        assert "revenue" in result.columns
        assert "product" in result.columns

    def test_no_id_column_returns_df_unchanged(self) -> None:
        """A DataFrame with no 'id' column is returned without modification."""
        df = pd.DataFrame({"revenue": [100.0], "product": ["Widget"]})
        result = _drop_db_cols(df)
        assert list(result.columns) == ["revenue", "product"]

    def test_input_df_not_mutated(self) -> None:
        """The original DataFrame is not modified in place."""
        df = pd.DataFrame({"id": [1], "revenue": [100.0]})
        original_cols = list(df.columns)
        _drop_db_cols(df)
        assert list(df.columns) == original_cols

    def test_empty_df_handled_without_error(self) -> None:
        """An empty DataFrame with an 'id' column is handled without error."""
        df = pd.DataFrame({"id": pd.Series([], dtype=int), "revenue": pd.Series([], dtype=float)})
        result = _drop_db_cols(df)
        assert "id" not in result.columns
        assert len(result) == 0


# ---------------------------------------------------------------------------
# _sheet_safe
# ---------------------------------------------------------------------------


class TestSheetSafe:
    """Tests for the private _sheet_safe helper."""

    def test_short_name_returned_unchanged(self) -> None:
        """A name under 31 characters is returned as-is."""
        assert _sheet_safe("East") == "East"

    def test_exactly_31_chars_returned_unchanged(self) -> None:
        """A name of exactly 31 characters is returned as-is."""
        name = "a" * 31
        assert _sheet_safe(name) == name

    def test_name_over_31_chars_truncated(self) -> None:
        """A name longer than 31 characters is truncated to 31."""
        name = "a" * 50
        result = _sheet_safe(name)
        assert len(result) == 31

    def test_truncation_preserves_leading_content(self) -> None:
        """Truncation keeps the first 31 characters, not the last."""
        name = "Clean - " + "x" * 40
        result = _sheet_safe(name)
        assert result == name[:31]

    def test_empty_string_returned_unchanged(self) -> None:
        """An empty string is returned as-is (length 0 ≤ 31)."""
        assert _sheet_safe("") == ""


# ---------------------------------------------------------------------------
# read_consolidated
# ---------------------------------------------------------------------------


class TestReadConsolidated:
    """Tests for read_consolidated."""

    def test_returns_dataframe(self, tmp_path: Path) -> None:
        """read_consolidated returns a pandas DataFrame."""
        db = _make_db(tmp_path, n_clean=1, n_quarantine=0)
        result = read_consolidated(db)
        assert isinstance(result, pd.DataFrame)

    def test_row_count_matches_db(self, tmp_path: Path) -> None:
        """Returned DataFrame has the same row count as the consolidated table."""
        db = _make_db(tmp_path, n_clean=5, n_quarantine=0)
        assert len(read_consolidated(db)) == 5

    def test_id_column_not_in_result(self, tmp_path: Path) -> None:
        """The 'id' column from SQLite is dropped before returning."""
        db = _make_db(tmp_path)
        result = read_consolidated(db)
        assert "id" not in result.columns

    def test_loaded_at_column_present(self, tmp_path: Path) -> None:
        """The 'loaded_at' timestamp column is included in the result."""
        db = _make_db(tmp_path)
        result = read_consolidated(db)
        assert "loaded_at" in result.columns

    def test_data_columns_present(self, tmp_path: Path) -> None:
        """Core data columns (date, product, region, revenue) are present."""
        db = _make_db(tmp_path)
        result = read_consolidated(db)
        for col in ("date", "product", "region", "revenue"):
            assert col in result.columns, f"Missing column: {col}"

    def test_empty_table_returns_empty_df(self, tmp_path: Path) -> None:
        """An empty consolidated table returns a zero-row DataFrame."""
        db = _make_db(tmp_path, n_clean=0, n_quarantine=0)
        result = read_consolidated(db)
        assert len(result) == 0

    def test_missing_db_raises_file_not_found(self, tmp_path: Path) -> None:
        """FileNotFoundError is raised when the database file does not exist."""
        with pytest.raises(FileNotFoundError, match="not found"):
            read_consolidated(tmp_path / "nonexistent.db")

    def test_accepts_string_path(self, tmp_path: Path) -> None:
        """A string path is accepted as well as a Path object."""
        db = _make_db(tmp_path)
        result = read_consolidated(str(db))
        assert isinstance(result, pd.DataFrame)


# ---------------------------------------------------------------------------
# read_quarantine
# ---------------------------------------------------------------------------


class TestReadQuarantine:
    """Tests for read_quarantine."""

    def test_returns_dataframe(self, tmp_path: Path) -> None:
        """read_quarantine returns a pandas DataFrame."""
        db = _make_db(tmp_path, n_quarantine=1)
        assert isinstance(read_quarantine(db), pd.DataFrame)

    def test_row_count_matches_db(self, tmp_path: Path) -> None:
        """Returned DataFrame has the same row count as the quarantine table."""
        db = _make_db(tmp_path, n_quarantine=4)
        assert len(read_quarantine(db)) == 4

    def test_id_column_not_in_result(self, tmp_path: Path) -> None:
        """The 'id' column from SQLite is dropped before returning."""
        db = _make_db(tmp_path, n_quarantine=1)
        assert "id" not in read_quarantine(db).columns

    def test_quarantine_reason_column_present(self, tmp_path: Path) -> None:
        """The 'quarantine_reason' column is present in the result."""
        db = _make_db(tmp_path, n_quarantine=1)
        assert "quarantine_reason" in read_quarantine(db).columns

    def test_quarantined_at_column_present(self, tmp_path: Path) -> None:
        """The 'quarantined_at' timestamp column is included in the result."""
        db = _make_db(tmp_path, n_quarantine=1)
        assert "quarantined_at" in read_quarantine(db).columns

    def test_empty_table_returns_empty_df(self, tmp_path: Path) -> None:
        """An empty quarantine table returns a zero-row DataFrame."""
        db = _make_db(tmp_path, n_quarantine=0)
        assert len(read_quarantine(db)) == 0

    def test_missing_db_raises_file_not_found(self, tmp_path: Path) -> None:
        """FileNotFoundError is raised when the database file does not exist."""
        with pytest.raises(FileNotFoundError, match="not found"):
            read_quarantine(tmp_path / "nonexistent.db")

    def test_accepts_string_path(self, tmp_path: Path) -> None:
        """A string path is accepted as well as a Path object."""
        db = _make_db(tmp_path, n_quarantine=1)
        assert isinstance(read_quarantine(str(db)), pd.DataFrame)


# ---------------------------------------------------------------------------
# split_by_column
# ---------------------------------------------------------------------------


class TestSplitByColumn:
    """Tests for split_by_column."""

    def _df(self) -> pd.DataFrame:
        """Return a three-row DataFrame with two distinct region values."""
        return pd.DataFrame({
            "region": ["East", "West", "East"],
            "revenue": [100.0, 200.0, 150.0],
        })

    def test_returns_dict(self) -> None:
        """split_by_column returns a dict."""
        assert isinstance(split_by_column(self._df(), "region"), dict)

    def test_one_key_per_unique_value(self) -> None:
        """The dict has one entry for each unique value in the column."""
        result = split_by_column(self._df(), "region")
        assert set(result.keys()) == {"East", "West"}

    def test_row_counts_sum_to_total(self) -> None:
        """Total rows across all slices equals the input DataFrame length."""
        result = split_by_column(self._df(), "region")
        total = sum(len(df) for df in result.values())
        assert total == len(self._df())

    def test_each_slice_contains_correct_rows(self) -> None:
        """Each slice contains only rows with the matching column value."""
        result = split_by_column(self._df(), "region")
        assert all(result["East"]["region"] == "East")
        assert all(result["West"]["region"] == "West")

    def test_grouping_column_retained_in_slice(self) -> None:
        """The grouping column is kept in each sub-DataFrame."""
        result = split_by_column(self._df(), "region")
        assert "region" in result["East"].columns

    def test_slice_indices_are_reset(self) -> None:
        """Each sub-DataFrame has a fresh integer index starting at 0."""
        result = split_by_column(self._df(), "region")
        assert list(result["East"].index) == list(range(len(result["East"])))

    def test_null_value_grouped_under_blank(self) -> None:
        """Rows with a null value in the group column appear under '(blank)'."""
        df = pd.DataFrame({"region": [None, "East"], "revenue": [100.0, 200.0]})
        result = split_by_column(df, "region")
        assert "(blank)" in result

    def test_empty_string_grouped_under_blank(self) -> None:
        """Rows with an empty string in the group column appear under '(blank)'."""
        df = pd.DataFrame({"region": ["", "East"], "revenue": [100.0, 200.0]})
        result = split_by_column(df, "region")
        assert "(blank)" in result

    def test_sheet_name_truncated_to_31_chars(self) -> None:
        """Keys longer than 31 characters are truncated to the Excel limit."""
        df = pd.DataFrame({"region": ["x" * 50], "revenue": [100.0]})
        result = split_by_column(df, "region")
        assert all(len(k) <= 31 for k in result.keys())

    def test_keys_sorted_alphabetically(self) -> None:
        """The dict keys are sorted by the unique column values."""
        df = pd.DataFrame({
            "region": ["West", "East", "Central"],
            "revenue": [1.0, 2.0, 3.0],
        })
        result = split_by_column(df, "region")
        assert list(result.keys()) == sorted(result.keys())

    def test_missing_column_raises_key_error(self) -> None:
        """KeyError is raised when the requested column is not in the DataFrame."""
        with pytest.raises(KeyError, match="not found"):
            split_by_column(self._df(), "nonexistent_col")

    def test_single_unique_value_returns_one_entry(self) -> None:
        """A column with one distinct value produces a dict with one entry."""
        df = pd.DataFrame({"region": ["East", "East", "East"], "revenue": [1.0, 2.0, 3.0]})
        result = split_by_column(df, "region")
        assert len(result) == 1


# ---------------------------------------------------------------------------
# build_sheet_map
# ---------------------------------------------------------------------------


class TestBuildSheetMap:
    """Tests for build_sheet_map."""

    def test_no_group_by_produces_clean_data_sheet(self) -> None:
        """Without group_by, the clean data sheet is named 'Clean Data'."""
        result = build_sheet_map(_make_clean_df(), _make_quarantine_df())
        assert "Clean Data" in result

    def test_no_group_by_produces_exactly_two_sheets(self) -> None:
        """Without group_by, the map contains exactly one clean sheet and quarantine."""
        result = build_sheet_map(_make_clean_df(), _make_quarantine_df())
        assert len(result) == 2

    def test_quarantine_sheet_always_present(self) -> None:
        """The 'Quarantine' sheet is always in the returned map."""
        result = build_sheet_map(_make_clean_df(), _make_quarantine_df())
        assert "Quarantine" in result

    def test_quarantine_sheet_is_last(self) -> None:
        """'Quarantine' is the final sheet in the ordered dict."""
        result = build_sheet_map(_make_clean_df(), _make_quarantine_df())
        assert list(result.keys())[-1] == "Quarantine"

    def test_quarantine_sheet_is_last_with_group_by(self) -> None:
        """'Quarantine' is still last when group_by splits clean data."""
        df = pd.DataFrame({
            "region": ["East", "West"], "revenue": [100.0, 200.0],
            "source_file": ["f.xlsx"] * 2, "source_row": [2, 3],
        })
        result = build_sheet_map(df, _make_quarantine_df(), group_by="region")
        assert list(result.keys())[-1] == "Quarantine"

    def test_group_by_produces_prefixed_sheet_names(self) -> None:
        """With group_by, clean sheets are prefixed with 'Clean - '."""
        df = pd.DataFrame({"region": ["East", "West"], "revenue": [100.0, 200.0]})
        result = build_sheet_map(df, _make_quarantine_df(), group_by="region")
        clean_keys = [k for k in result if k != "Quarantine"]
        assert all(k.startswith("Clean - ") for k in clean_keys)

    def test_group_by_produces_one_sheet_per_unique_value(self) -> None:
        """With group_by, there is one clean sheet per distinct value in the column."""
        df = pd.DataFrame({"region": ["East", "West", "East"], "revenue": [1.0, 2.0, 3.0]})
        result = build_sheet_map(df, _make_quarantine_df(), group_by="region")
        clean_keys = [k for k in result if k != "Quarantine"]
        assert len(clean_keys) == 2

    def test_clean_data_df_placed_in_clean_sheet(self) -> None:
        """The clean DataFrame is the value behind the 'Clean Data' key."""
        clean = _make_clean_df(3)
        result = build_sheet_map(clean, _make_quarantine_df())
        assert len(result["Clean Data"]) == 3

    def test_quarantine_df_placed_in_quarantine_sheet(self) -> None:
        """The quarantine DataFrame is the value behind the 'Quarantine' key."""
        quarantine = _make_quarantine_df(2)
        result = build_sheet_map(_make_clean_df(), quarantine)
        assert len(result["Quarantine"]) == 2

    def test_empty_clean_df_handled_gracefully(self) -> None:
        """An empty clean DataFrame produces a map without error."""
        result = build_sheet_map(pd.DataFrame(), _make_quarantine_df())
        assert "Clean Data" in result
        assert len(result["Clean Data"]) == 0

    def test_empty_quarantine_df_handled_gracefully(self) -> None:
        """An empty quarantine DataFrame produces a map without error."""
        result = build_sheet_map(_make_clean_df(), pd.DataFrame())
        assert "Quarantine" in result
        assert len(result["Quarantine"]) == 0


# ---------------------------------------------------------------------------
# write_workbook
# ---------------------------------------------------------------------------


class TestWriteWorkbook:
    """Tests for write_workbook, including indirect tests of _autosize_columns
    and _freeze_header via the written .xlsx file."""

    def _simple_map(self) -> dict[str, pd.DataFrame]:
        """Return a minimal two-sheet map for use in tests."""
        return build_sheet_map(_make_clean_df(3), _make_quarantine_df(1))

    def test_creates_xlsx_file_on_disk(self, tmp_path: Path) -> None:
        """write_workbook creates the .xlsx file at the specified path."""
        out = tmp_path / "out.xlsx"
        write_workbook(self._simple_map(), out)
        assert out.exists()

    def test_returns_export_result_instance(self, tmp_path: Path) -> None:
        """write_workbook returns an ExportResult dataclass."""
        result = write_workbook(self._simple_map(), tmp_path / "out.xlsx")
        assert isinstance(result, ExportResult)

    def test_result_output_path_matches_argument(self, tmp_path: Path) -> None:
        """ExportResult.output_path equals the path passed to write_workbook."""
        out = tmp_path / "out.xlsx"
        result = write_workbook(self._simple_map(), out)
        assert result.output_path == out

    def test_result_sheets_match_sheet_map_keys(self, tmp_path: Path) -> None:
        """ExportResult.sheets lists the same keys as the sheet_map, in order."""
        sheet_map = self._simple_map()
        result = write_workbook(sheet_map, tmp_path / "out.xlsx")
        assert result.sheets == list(sheet_map.keys())

    def test_n_clean_counts_non_quarantine_rows(self, tmp_path: Path) -> None:
        """n_clean equals the total rows across all non-Quarantine sheets."""
        result = write_workbook(
            build_sheet_map(_make_clean_df(5), _make_quarantine_df(2)),
            tmp_path / "out.xlsx",
        )
        assert result.n_clean == 5

    def test_n_quarantine_counts_quarantine_sheet_rows(self, tmp_path: Path) -> None:
        """n_quarantine equals the row count of the Quarantine sheet."""
        result = write_workbook(
            build_sheet_map(_make_clean_df(5), _make_quarantine_df(2)),
            tmp_path / "out.xlsx",
        )
        assert result.n_quarantine == 2

    def test_n_clean_sums_across_multiple_clean_sheets(self, tmp_path: Path) -> None:
        """n_clean accumulates rows from multiple clean sheets when group_by is used."""
        df = pd.DataFrame({
            "region": ["East"] * 3 + ["West"] * 2,
            "revenue": [100.0] * 5,
        })
        sheet_map = build_sheet_map(df, _make_quarantine_df(1), group_by="region")
        result = write_workbook(sheet_map, tmp_path / "out.xlsx")
        assert result.n_clean == 5

    def test_creates_missing_parent_directories(self, tmp_path: Path) -> None:
        """write_workbook creates any missing directories for the output path."""
        nested = tmp_path / "a" / "b" / "out.xlsx"
        write_workbook(self._simple_map(), nested)
        assert nested.exists()

    def test_workbook_has_correct_sheet_names(self, tmp_path: Path) -> None:
        """The written workbook contains exactly the expected sheet names."""
        out = tmp_path / "out.xlsx"
        write_workbook(self._simple_map(), out)
        wb = openpyxl.load_workbook(out)
        assert set(wb.sheetnames) == {"Clean Data", "Quarantine"}

    def test_header_row_frozen_on_clean_sheet(self, tmp_path: Path) -> None:
        """The first row is frozen (freeze_panes='A2') on the Clean Data sheet."""
        out = tmp_path / "out.xlsx"
        write_workbook(self._simple_map(), out)
        wb = openpyxl.load_workbook(out)
        assert wb["Clean Data"].freeze_panes == "A2"

    def test_header_row_frozen_on_quarantine_sheet(self, tmp_path: Path) -> None:
        """The first row is frozen on the Quarantine sheet."""
        out = tmp_path / "out.xlsx"
        write_workbook(self._simple_map(), out)
        wb = openpyxl.load_workbook(out)
        assert wb["Quarantine"].freeze_panes == "A2"

    def test_column_widths_set_to_positive_values(self, tmp_path: Path) -> None:
        """All column widths on the Clean Data sheet are set to a positive value."""
        out = tmp_path / "out.xlsx"
        write_workbook(self._simple_map(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Clean Data"]
        widths = [
            dim.width
            for dim in ws.column_dimensions.values()
            if dim.width is not None
        ]
        assert len(widths) > 0
        assert all(w > 0 for w in widths)

    def test_empty_quarantine_sheet_written_without_error(self, tmp_path: Path) -> None:
        """An empty quarantine DataFrame produces a valid (header-only) sheet."""
        out = tmp_path / "out.xlsx"
        result = write_workbook(
            build_sheet_map(_make_clean_df(2), pd.DataFrame()),
            out,
        )
        assert result.n_quarantine == 0
        assert out.exists()

    def test_accepts_string_output_path(self, tmp_path: Path) -> None:
        """A string output path is accepted in addition to a Path object."""
        out = str(tmp_path / "out.xlsx")
        result = write_workbook(self._simple_map(), out)
        assert Path(out).exists()
        assert isinstance(result, ExportResult)

    def test_data_rows_written_to_workbook(self, tmp_path: Path) -> None:
        """Row data from the DataFrames is present in the written workbook."""
        out = tmp_path / "out.xlsx"
        write_workbook(self._simple_map(), out)
        wb = openpyxl.load_workbook(out)
        ws = wb["Clean Data"]
        # Row 1 is header; row 2 is first data row
        assert ws.max_row >= 2


# ---------------------------------------------------------------------------
# build_summary
# ---------------------------------------------------------------------------


class TestBuildSummary:
    """Tests for build_summary."""

    def test_contains_output_path(self) -> None:
        """Summary string includes the workbook file path."""
        result = _make_result(output_path=Path("data/output/out.xlsx"))
        assert "data/output/out.xlsx" in build_summary(result)

    def test_contains_n_clean(self) -> None:
        """Summary string includes the clean row count."""
        assert "42" in build_summary(_make_result(n_clean=42))

    def test_contains_n_quarantine(self) -> None:
        """Summary string includes the quarantine row count."""
        assert "7" in build_summary(_make_result(n_quarantine=7))

    def test_contains_sheet_names(self) -> None:
        """Summary string lists each sheet name."""
        result = _make_result(sheets=["Clean Data", "Quarantine"])
        summary = build_summary(result)
        assert "Clean Data" in summary
        assert "Quarantine" in summary

    def test_zero_counts_shown_not_omitted(self) -> None:
        """Zero row counts are explicitly present in the summary, not omitted."""
        summary = build_summary(_make_result(n_clean=0, n_quarantine=0))
        assert "0" in summary

    def test_returns_string(self) -> None:
        """build_summary returns a str."""
        assert isinstance(build_summary(_make_result()), str)

    def test_multiple_clean_sheets_all_listed(self) -> None:
        """All sheet names appear in the summary when group_by produces multiple sheets."""
        result = _make_result(
            sheets=["Clean - East", "Clean - West", "Quarantine"],
            n_clean=5,
            n_quarantine=1,
        )
        summary = build_summary(result)
        assert "Clean - East" in summary
        assert "Clean - West" in summary


# ---------------------------------------------------------------------------
# export  (integration)
# ---------------------------------------------------------------------------


class TestExport:
    """Integration tests for the export pipeline orchestrator."""

    def test_creates_xlsx_file(self, tmp_path: Path) -> None:
        """export creates a .xlsx file on disk."""
        db = _make_db(tmp_path, n_clean=3, n_quarantine=1)
        out = tmp_path / "out.xlsx"
        export(db, out)
        assert out.exists()

    def test_returns_export_result(self, tmp_path: Path) -> None:
        """export returns an ExportResult."""
        db = _make_db(tmp_path, n_clean=2, n_quarantine=1)
        result = export(db, tmp_path / "out.xlsx")
        assert isinstance(result, ExportResult)

    def test_n_clean_matches_db_row_count(self, tmp_path: Path) -> None:
        """n_clean equals the number of rows in the consolidated table."""
        db = _make_db(tmp_path, n_clean=4, n_quarantine=0)
        result = export(db, tmp_path / "out.xlsx")
        assert result.n_clean == 4

    def test_n_quarantine_matches_db_row_count(self, tmp_path: Path) -> None:
        """n_quarantine equals the number of rows in the quarantine table."""
        db = _make_db(tmp_path, n_clean=0, n_quarantine=3)
        result = export(db, tmp_path / "out.xlsx")
        assert result.n_quarantine == 3

    def test_default_produces_two_sheets(self, tmp_path: Path) -> None:
        """Without group_by, the workbook has exactly two sheets."""
        db = _make_db(tmp_path)
        result = export(db, tmp_path / "out.xlsx")
        assert len(result.sheets) == 2

    def test_group_by_splits_clean_data(self, tmp_path: Path) -> None:
        """group_by produces one clean sheet per distinct region value."""
        db = tmp_path / "multi.db"
        with sqlite3.connect(db) as conn:
            conn.executescript("""
                CREATE TABLE consolidated (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    region TEXT,
                    revenue REAL,
                    source_file TEXT,
                    source_row INTEGER,
                    date TEXT,
                    product TEXT,
                    sales_rep TEXT,
                    customer TEXT,
                    quantity REAL,
                    loaded_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ', 'now'))
                );
                CREATE TABLE quarantine (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    quarantine_reason TEXT NOT NULL,
                    source_file TEXT,
                    source_row INTEGER,
                    date TEXT, product TEXT, region TEXT,
                    sales_rep TEXT, customer TEXT, quantity TEXT, revenue TEXT,
                    quarantined_at TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ', 'now'))
                );
            """)
            conn.execute(
                "INSERT INTO consolidated (region, revenue, source_file, source_row, date, product, sales_rep, customer, quantity) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("East", 100.0, "f.xlsx", 2, "2024-01-01", "W", "Alice", "Corp", 1.0),
            )
            conn.execute(
                "INSERT INTO consolidated (region, revenue, source_file, source_row, date, product, sales_rep, customer, quantity) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("West", 200.0, "f.xlsx", 3, "2024-01-01", "W", "Bob", "Corp", 2.0),
            )
            conn.commit()
        result = export(db, tmp_path / "out.xlsx", group_by="region")
        clean_sheets = [s for s in result.sheets if s != "Quarantine"]
        assert len(clean_sheets) == 2

    def test_missing_db_raises_file_not_found(self, tmp_path: Path) -> None:
        """FileNotFoundError is raised when the database file does not exist."""
        with pytest.raises(FileNotFoundError):
            export(tmp_path / "missing.db", tmp_path / "out.xlsx")

    def test_creates_missing_output_directories(self, tmp_path: Path) -> None:
        """export creates any missing parent directories for the output path."""
        db = _make_db(tmp_path)
        nested_out = tmp_path / "subdir" / "deep" / "out.xlsx"
        export(db, nested_out)
        assert nested_out.exists()

    def test_quarantine_sheet_is_always_last(self, tmp_path: Path) -> None:
        """The Quarantine sheet is the last tab in the workbook."""
        db = _make_db(tmp_path, n_clean=3, n_quarantine=2)
        result = export(db, tmp_path / "out.xlsx")
        assert result.sheets[-1] == "Quarantine"

    def test_empty_db_tables_produce_valid_workbook(self, tmp_path: Path) -> None:
        """A database with zero rows in both tables produces a valid workbook."""
        db = _make_db(tmp_path, n_clean=0, n_quarantine=0)
        result = export(db, tmp_path / "out.xlsx")
        assert result.n_clean == 0
        assert result.n_quarantine == 0
        assert (tmp_path / "out.xlsx").exists()

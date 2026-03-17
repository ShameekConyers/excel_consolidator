"""
Generate sample Excel and CSV files for the excel_consolidator demo.
Run once: python generate_samples.py

All files represent quarterly sales transactions — same domain, different formatting chaos.
Row counts are sampled from Poisson(lambda=50). Fixed seed (42) for reproducibility.
"""

import csv
import os
import random

import numpy as np
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font

SEED = 42
random.seed(SEED)
np.random.seed(SEED)

OUT = os.path.dirname(os.path.abspath(__file__))

PRODUCTS = ["Analytics Pro", "Data Suite", "Reporting Pkg", "Insights Hub", "DataViz Pro"]
UNIT_PRICES = {
    "Analytics Pro": 1500,
    "Data Suite": 2200,
    "Reporting Pkg": 1450,
    "Insights Hub": 1800,
    "DataViz Pro": 2500,
}
REPS = {"West": "Dana Lee", "East": "Tom Rivera", "Central": "Sara Kim"}
REPS_SHORT = {"West": "D. Lee", "East": "T. Rivera", "Central": "S. Kim"}
CUSTOMERS = [
    "Acme Corp", "Globex LLC", "Initech", "Umbrella Inc",
    "Soylent Corp", "Massive Dynamics", "Wayne Enterprises", "Oscorp",
]
MONTHS_FULL = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
MONTHS_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def rand_date(start: date, end: date) -> date:
    """Return a uniformly random date in the inclusive range [start, end].

    Args:
        start (date): Earliest possible date.
        end (date): Latest possible date.

    Returns:
        date: A randomly selected date between start and end.
    """
    delta = max((end - start).days, 1)
    return start + timedelta(days=random.randint(0, delta))


def poisson_n(lam: int = 50) -> int:
    """Sample a row count from a Poisson distribution.

    Args:
        lam (int): Lambda parameter — the expected mean row count. Defaults to 50.

    Returns:
        int: Sampled row count, floored at 8 to avoid degenerate files.
    """
    return max(int(np.random.poisson(lam)), 8)


def fmt_iso(d: date) -> str:
    """Format a date as ISO 8601.

    Args:
        d (date): The date to format.

    Returns:
        str: Date string in YYYY-MM-DD format, e.g. '2024-03-15'.
    """
    return d.strftime("%Y-%m-%d")


def fmt_us(d: date) -> str:
    """Format a date in US locale order.

    Args:
        d (date): The date to format.

    Returns:
        str: Date string in MM/DD/YYYY format, e.g. '03/15/2024'.
    """
    return d.strftime("%m/%d/%Y")


def fmt_written(d: date) -> str:
    """Format a date with the full written month name.

    Args:
        d (date): The date to format.

    Returns:
        str: Date string e.g. 'March 15, 2024'.
    """
    return f"{MONTHS_FULL[d.month - 1]} {d.day}, {d.year}"


def fmt_written_short(d: date) -> str:
    """Format a date with an abbreviated three-letter month name.

    Args:
        d (date): The date to format.

    Returns:
        str: Date string e.g. 'Mar 15, 2024'.
    """
    return f"{MONTHS_SHORT[d.month - 1]} {d.day}, {d.year}"


def fmt_mixed_q2(d: date) -> str:
    """Format a date using a randomly chosen style to simulate Q2's mixed-format chaos.

    Probabilities: US format (40%), ISO (35%), full written month (25%).

    Args:
        d (date): The date to format.

    Returns:
        str: Date string in one of three formats chosen at random.
    """
    r = random.random()
    if r < 0.40:
        return fmt_us(d)
    elif r < 0.75:
        return fmt_iso(d)
    else:
        return fmt_written(d)


def fmt_written_q4(d: date) -> str:
    """Format a date in written style for Q4, mixing full and abbreviated month names.

    Probabilities: full month name (60%), abbreviated month name (40%).

    Args:
        d (date): The date to format.

    Returns:
        str: Date string in 'Month D, YYYY' or 'Mon D, YYYY' format.
    """
    return fmt_written(d) if random.random() < 0.6 else fmt_written_short(d)


def insert_bad_rows(clean_rows: list, bad_rows: list) -> list:
    """Insert each bad row at a random position within a list of clean rows.

    Each bad row is inserted independently, so bad rows may end up adjacent
    or spread across the file depending on the random draw.

    Args:
        clean_rows (list): List of well-formed data rows.
        bad_rows (list): List of intentionally malformed rows to scatter in.

    Returns:
        list: Combined list with bad rows inserted at random positions.
    """
    result = list(clean_rows)
    for bad in bad_rows:
        pos = random.randint(0, len(result))
        result.insert(pos, bad)
    return result


# ---------------------------------------------------------------------------
# Q1_2024_sales.xlsx
# Columns: Date | Product | Region | Sales Rep | Customer | Qty | Revenue
# Clean column names, ISO dates. Bad: negative revenue, missing customer, empty row.
# ---------------------------------------------------------------------------
def make_q1() -> None:
    """Generate Q1_2024_sales.xlsx — the clean-baseline quarterly sales file.

    Uses standard column names and ISO dates. Row count is sampled from
    Poisson(50). Three bad rows are seeded in: one negative revenue, one
    missing customer, and one entirely empty row.

    Args:
        None

    Returns:
        None: Writes Q1_2024_sales.xlsx to the OUT directory.
    """
    n = poisson_n()
    headers = ["Date", "Product", "Region", "Sales Rep", "Customer", "Qty", "Revenue"]
    start, end = date(2024, 1, 1), date(2024, 3, 31)

    bad_rows = [
        ["2024-03-18", "Analytics Pro", "Central", "Sara Kim", "Acme Corp", 1, -450.00],  # negative revenue
        ["2024-03-21", "Reporting Pkg", "West", "Dana Lee", None, 3, 4350.00],             # missing customer
        [None, None, None, None, None, None, None],                                         # entirely empty row
    ]
    clean = []
    for _ in range(n - len(bad_rows)):
        region = random.choice(["West", "East", "Central"])
        product = random.choice(PRODUCTS)
        qty = random.randint(1, 8)
        revenue = round(qty * UNIT_PRICES[product] * random.uniform(0.9, 1.1), 2)
        d = rand_date(start, end)
        clean.append([fmt_iso(d), product, region, REPS[region], random.choice(CUSTOMERS), qty, revenue])
    clean.sort(key=lambda x: x[0])
    rows = insert_bad_rows(clean, bad_rows)

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # wb.active is Optional[Worksheet] in openpyxl stubs; assert narrows the type for Pylance
    ws.title = "Q1 Sales"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(os.path.join(OUT, "Q1_2024_sales.xlsx"))
    print(f"Q1_2024_sales.xlsx:      {len(rows):>3} rows  ({len(bad_rows)} bad)")


# ---------------------------------------------------------------------------
# Q2_2024_sales.xlsx
# Columns: transaction_date | product_name | territory | rep | client | units | Rev.
# Renamed cols, abbreviated rep names, mixed date formats.
# Bad: comment row, "pending" in revenue.
# ---------------------------------------------------------------------------
def make_q2() -> None:
    """Generate Q2_2024_sales.xlsx — renamed columns, mixed date formats, abbreviated rep names.

    Every column has a different name from Q1 to simulate real-world schema drift.
    Date formats are randomly assigned per row (US, ISO, or written). Rep names
    are abbreviated (e.g. 'D. Lee'). Two bad rows are seeded: one comment row
    left in the date field and one non-numeric revenue value ('pending').

    Args:
        None

    Returns:
        None: Writes Q2_2024_sales.xlsx to the OUT directory.
    """
    n = poisson_n()
    headers = ["transaction_date", "product_name", "territory", "rep", "client", "units", "Rev."]
    start, end = date(2024, 4, 1), date(2024, 6, 30)

    bad_rows = [
        ["Mike please update these numbers before EOD", None, None, None, None, None, None],  # comment row
        ["06/17/2024", "Analytics Pro", "Central", "S. Kim", "Acme Corp", 2, "pending"],      # text in revenue
    ]
    clean = []
    for _ in range(n - len(bad_rows)):
        region = random.choice(["West", "East", "Central"])
        product = random.choice(PRODUCTS)
        qty = random.randint(1, 8)
        revenue = round(qty * UNIT_PRICES[product] * random.uniform(0.9, 1.1), 2)
        d = rand_date(start, end)
        clean.append([fmt_mixed_q2(d), product, region, REPS_SHORT[region],
                       random.choice(CUSTOMERS), qty, revenue])
    rows = insert_bad_rows(clean, bad_rows)

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # wb.active is Optional[Worksheet] in openpyxl stubs; assert narrows the type for Pylance
    ws.title = "Sheet1"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(os.path.join(OUT, "Q2_2024_sales.xlsx"))
    print(f"Q2_2024_sales.xlsx:      {len(rows):>3} rows  ({len(bad_rows)} bad)")


# ---------------------------------------------------------------------------
# Q3_2024_sales.xlsx
# Columns: Date | Item | Area | Salesperson | Account | Quantity | Total Revenue
# Further-renamed columns. Bad: impossible date (month 13), zero qty, "TBD" in qty,
# negative revenue, impossible date (Feb 30).
# ---------------------------------------------------------------------------
def make_q3() -> None:
    """Generate Q3_2024_sales.xlsx — most heavily renamed columns and highest bad-row count.

    Column names diverge furthest from the canonical schema (e.g. 'Item', 'Area',
    'Account'). Five bad rows cover the widest variety of error types: two
    impossible dates, a zero-quantity row, a non-numeric quantity ('TBD'), and
    a negative revenue.

    Args:
        None

    Returns:
        None: Writes Q3_2024_sales.xlsx to the OUT directory.
    """
    n = poisson_n()
    headers = ["Date", "Item", "Area", "Salesperson", "Account", "Quantity", "Total Revenue"]
    start, end = date(2024, 7, 1), date(2024, 9, 30)

    bad_rows = [
        ["2024-13-01", "Reporting Pkg", "Central", "Sara Kim", "Initech", 3, 4350.00],           # month 13 — impossible date
        ["2024-08-19", "Reporting Pkg", "West", "Dana Lee", "Acme Corp", 0, 0.00],               # zero quantity
        ["2024-09-02", "Analytics Pro", "Central", "Sara Kim", "Umbrella Inc", "TBD", 3000.00],  # text in quantity
        ["2024-09-16", "Data Suite", "East", "Tom Rivera", "Initech", 2, -880.00],               # negative revenue
        ["2024-02-30", "Reporting Pkg", "East", "Tom Rivera", "Globex LLC", 2, 2900.00],         # Feb 30 — impossible date
    ]
    clean = []
    for _ in range(n - len(bad_rows)):
        region = random.choice(["West", "East", "Central"])
        product = random.choice(PRODUCTS)
        qty = random.randint(1, 8)
        revenue = round(qty * UNIT_PRICES[product] * random.uniform(0.9, 1.1), 2)
        d = rand_date(start, end)
        clean.append([fmt_iso(d), product, region, REPS[region], random.choice(CUSTOMERS), qty, revenue])
    clean.sort(key=lambda x: x[0])
    rows = insert_bad_rows(clean, bad_rows)

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # wb.active is Optional[Worksheet] in openpyxl stubs; assert narrows the type for Pylance
    ws.title = "Q3"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(os.path.join(OUT, "Q3_2024_sales.xlsx"))
    print(f"Q3_2024_sales.xlsx:      {len(rows):>3} rows  ({len(bad_rows)} bad)")


# ---------------------------------------------------------------------------
# Q4_2024_sales.xlsx
# Columns: date | product | region | sales_rep | customer | quantity | amount
# Written-out month names throughout. Bad: far-future date (2099), duplicate row,
# "N/A" in amount.
# ---------------------------------------------------------------------------
def make_q4() -> None:
    """Generate Q4_2024_sales.xlsx — written month names throughout, duplicate row, future-date outlier.

    Every clean date uses a written month format ('October 1, 2024' or 'Oct 1, 2024'),
    making the column non-trivially parseable. Four bad rows: a far-future date
    (2099), two identical rows to form an exact duplicate, and a non-numeric
    amount ('N/A').

    Args:
        None

    Returns:
        None: Writes Q4_2024_sales.xlsx to the OUT directory.
    """
    n = poisson_n()
    headers = ["date", "product", "region", "sales_rep", "customer", "quantity", "amount"]
    start, end = date(2024, 10, 1), date(2024, 12, 31)

    dup_row = ["November 19, 2024", "Data Suite", "East", "Tom Rivera", "Globex LLC", 1, 2200.00]
    bad_rows = [
        ["2099-01-01", "Analytics Pro", "Central", "Sara Kim", "Initech", 2, 3000.00],       # far-future date
        list(dup_row),                                                                         # first occurrence
        list(dup_row),                                                                         # exact duplicate
        ["December 10, 2024", "Data Suite", "East", "Tom Rivera", "Initech", 2, "N/A"],      # text in amount
    ]
    clean = []
    for _ in range(n - len(bad_rows)):
        region = random.choice(["West", "East", "Central"])
        product = random.choice(PRODUCTS)
        qty = random.randint(1, 8)
        revenue = round(qty * UNIT_PRICES[product] * random.uniform(0.9, 1.1), 2)
        d = rand_date(start, end)
        clean.append([fmt_written_q4(d), product, region, REPS[region],
                       random.choice(CUSTOMERS), qty, revenue])
    rows = insert_bad_rows(clean, bad_rows)

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # wb.active is Optional[Worksheet] in openpyxl stubs; assert narrows the type for Pylance
    ws.title = "Q4 Data"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(os.path.join(OUT, "Q4_2024_sales.xlsx"))
    print(f"Q4_2024_sales.xlsx:      {len(rows):>3} rows  ({len(bad_rows)} bad)")


# ---------------------------------------------------------------------------
# west_region_2024.xlsx
# Columns: Transaction Date | Product Line | Rep Name | Client Name | Units Sold | Revenue ($)
# West-only (no Region column). Extra title row + blank spacer at top.
# Bad: negative revenue, comment row.
# ---------------------------------------------------------------------------
def make_west() -> None:
    """Generate west_region_2024.xlsx — West-only data with a title row above the headers.

    No Region column (all rows are West by definition). Row 1 is a plain-text
    title, row 2 is blank, and row 3 is the actual bold header row — the
    consolidator must detect and skip the two non-data rows before reading.
    Two bad rows: one negative revenue and one comment left in the date field.

    Args:
        None

    Returns:
        None: Writes west_region_2024.xlsx to the OUT directory.
    """
    n = poisson_n()
    headers = ["Transaction Date", "Product Line", "Rep Name", "Client Name", "Units Sold", "Revenue ($)"]
    start, end = date(2024, 1, 1), date(2024, 12, 31)

    bad_rows = [
        ["2024-06-20", "Data Suite", "Dana Lee", "Globex LLC", 1, -2200.00],                       # negative revenue
        ["TODO: Dana confirm Q4 numbers with finance", None, None, None, None, None],               # comment row
    ]
    clean = []
    for _ in range(n - len(bad_rows)):
        product = random.choice(PRODUCTS)
        qty = random.randint(1, 8)
        revenue = round(qty * UNIT_PRICES[product] * random.uniform(0.9, 1.1), 2)
        d = rand_date(start, end)
        clean.append([fmt_iso(d), product, "Dana Lee", random.choice(CUSTOMERS), qty, revenue])
    clean.sort(key=lambda x: x[0])
    rows = insert_bad_rows(clean, bad_rows)

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # wb.active is Optional[Worksheet] in openpyxl stubs; assert narrows the type for Pylance
    ws.title = "West Region"
    ws.append(["West Region Sales Summary — FY 2024"])
    ws.append([])  # blank spacer row
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[3]:
        cell.font = bold
    for r in rows:
        ws.append(r)
    wb.save(os.path.join(OUT, "west_region_2024.xlsx"))
    print(f"west_region_2024.xlsx:   {len(rows):>3} rows  ({len(bad_rows)} bad)  + 2 title rows")


# ---------------------------------------------------------------------------
# east_central_Q2_Q3.xlsx
# Columns: Sale Date | Product | Region | Rep | Customer | Qty | $
# East + Central only, Q2–Q3. Short header names.
# Bad: section-divider row, currency string in revenue, missing rep+customer.
# ---------------------------------------------------------------------------
def make_east_central() -> None:
    """Generate east_central_Q2_Q3.xlsx — East and Central regions, Q2–Q3, short header names.

    Uses terse column names including '$' for revenue. A section-divider row
    ('--- Q3 below ---') appears mid-file as if someone manually annotated the
    spreadsheet. Three bad rows: the divider, a currency-formatted string
    ('$3,000') in the revenue column, and a row with both rep and customer missing.

    Args:
        None

    Returns:
        None: Writes east_central_Q2_Q3.xlsx to the OUT directory.
    """
    n = poisson_n()
    headers = ["Sale Date", "Product", "Region", "Rep", "Customer", "Qty", "$"]
    start, end = date(2024, 4, 1), date(2024, 9, 30)

    bad_rows = [
        ["--- Q3 below ---", None, None, None, None, None, None],            # section-divider row
        ["2024-09-03", "Analytics Pro", "Central", "Sara Kim", "Acme Corp", 2, "$3,000"],  # currency string in revenue
        ["2024-09-27", "Reporting Pkg", "Central", None, None, 4, 5800.00],  # missing rep and customer
    ]
    clean = []
    for _ in range(n - len(bad_rows)):
        region = random.choice(["East", "Central"])
        product = random.choice(PRODUCTS)
        qty = random.randint(1, 8)
        revenue = round(qty * UNIT_PRICES[product] * random.uniform(0.9, 1.1), 2)
        d = rand_date(start, end)
        clean.append([fmt_iso(d), product, region, REPS[region], random.choice(CUSTOMERS), qty, revenue])
    clean.sort(key=lambda x: x[0])
    rows = insert_bad_rows(clean, bad_rows)

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # wb.active is Optional[Worksheet] in openpyxl stubs; assert narrows the type for Pylance
    ws.title = "Sales"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(os.path.join(OUT, "east_central_Q2_Q3.xlsx"))
    print(f"east_central_Q2_Q3.xlsx: {len(rows):>3} rows  ({len(bad_rows)} bad)")


# ---------------------------------------------------------------------------
# pipeline_q3_q4.csv
# Clean-ish CSV. Standard headers, ISO dates, all three regions.
# Bad: negative revenue, "unknown" in quantity.
# ---------------------------------------------------------------------------
def make_pipeline_csv() -> None:
    """Generate pipeline_q3_q4.csv — the clean-baseline CSV file spanning Q3–Q4.

    Uses canonical snake_case column names and ISO dates throughout, making it
    the easiest file for the consolidator to ingest. Two bad rows are seeded:
    one negative revenue and one non-numeric quantity ('unknown').

    Args:
        None

    Returns:
        None: Writes pipeline_q3_q4.csv to the OUT directory.
    """
    n = poisson_n()
    headers = ["date", "product", "region", "sales_rep", "customer", "quantity", "revenue"]
    start, end = date(2024, 7, 1), date(2024, 12, 31)

    bad_rows = [
        ["2024-09-11", "Analytics Pro", "Central", "Sara Kim", "Initech", 1, -1500.00],        # negative revenue
        ["2024-11-14", "Data Suite", "East", "Tom Rivera", "Initech", "unknown", 4400.00],      # text in quantity
    ]
    clean = []
    for _ in range(n - len(bad_rows)):
        region = random.choice(["West", "East", "Central"])
        product = random.choice(PRODUCTS)
        qty = random.randint(1, 8)
        revenue = round(qty * UNIT_PRICES[product] * random.uniform(0.9, 1.1), 2)
        d = rand_date(start, end)
        clean.append([fmt_iso(d), product, region, REPS[region], random.choice(CUSTOMERS), qty, revenue])
    clean.sort(key=lambda x: x[0])
    rows = insert_bad_rows(clean, bad_rows)

    path = os.path.join(OUT, "pipeline_q3_q4.csv")
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"pipeline_q3_q4.csv:      {len(rows):>3} rows  ({len(bad_rows)} bad)")


# ---------------------------------------------------------------------------
# returns_flagged.csv
# Return/refund records. Negative refund amounts are EXPECTED (they're returns).
# Bad: comment row, empty refund amount, impossible date.
# ---------------------------------------------------------------------------
def make_returns_csv() -> None:
    """Generate returns_flagged.csv — return and refund records where negative amounts are valid.

    All clean rows carry a negative Refund Amt by design (they are product returns,
    not sales). This creates an intentional ambiguity: the consolidator's validation
    rules will flag negative revenue unless the file type is handled separately.
    US date format (MM/DD/YYYY) throughout. Three bad rows: a comment row, a row
    with an empty refund amount, and a row with an impossible date (month 13).

    Args:
        None

    Returns:
        None: Writes returns_flagged.csv to the OUT directory.
    """
    n = poisson_n()
    headers = ["Return Date", "Product Name", "Region", "Sales Rep", "Customer", "Units", "Refund Amt"]
    start, end = date(2024, 1, 1), date(2024, 9, 30)

    bad_rows = [
        ["Finance: do not process until dispute resolved", "", "", "", "", "", ""],  # comment row
        ["07/14/2024", "Data Suite", "East", "Tom Rivera", "Initech", 1, ""],       # empty refund amount
        ["13/01/2024", "Analytics Pro", "East", "Tom Rivera", "Soylent Corp", 1, -1500.00],  # month 13 — impossible date
    ]
    clean = []
    for _ in range(n - len(bad_rows)):
        region = random.choice(["West", "East", "Central"])
        product = random.choice(PRODUCTS)
        qty = random.randint(1, 3)
        refund = -round(qty * UNIT_PRICES[product] * random.uniform(0.9, 1.0), 2)
        d = rand_date(start, end)
        clean.append([fmt_us(d), product, region, REPS[region], random.choice(CUSTOMERS), qty, refund])
    rows = insert_bad_rows(clean, bad_rows)

    path = os.path.join(OUT, "returns_flagged.csv")
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"returns_flagged.csv:     {len(rows):>3} rows  ({len(bad_rows)} bad)")


if __name__ == "__main__":
    print("Generating sample files (Poisson lambda=50, seed=42)...\n")
    make_q1()
    make_q2()
    make_q3()
    make_q4()
    make_west()
    make_east_central()
    make_pipeline_csv()
    make_returns_csv()
    print(f"\nDone. 6 Excel + 2 CSV files written to {OUT}")
